from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from users.models import Usuario, Rol, PasswordResetToken, ControlIP, generar_codigo_6_digitos
from audit.models import Auditoria
from django.shortcuts import get_object_or_404
from .models import Rol, Funcion, FuncionRol
from modules.models import Modulo
from django.core.mail import send_mail
from django.urls import reverse
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q
from .schema import validate_microsoft_token, generate_jwt
import json
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import os
import jwt as pyjwt
# ======================================================
# API Keys válidas por módulo externo (El Guardia)
# En producción móvelas a variables de entorno o la BD.
# ======================================================
API_KEYS_VALIDAS = {
    "dev_key_cxc_111":         "CXC",
    "dev_key_facturacion_222": "FACTURACION",
    "dev_key_inventario_333":  "INVENTARIO",
    "dev_key_compras_444":     "COMPRAS",
    "dev_key_seguridad_555":   "SEGURIDAD",
}
JWT_SECRET = os.getenv('SECRET_KEY', 'secret')
@csrf_exempt
@require_POST
def api_auth_login(request):
    """API REST de Autenticación (El Guardia).
    Recibe JSON: {api_key, usuario, clave, ip}
    Aplica lógica de bloqueo por fuerza bruta y,
    si todo está bien, devuelve un JWT.
    """
    # -- 0. Parsear body JSON ------------------------------------------------
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return _json_response(400, False, "El cuerpo de la petición no es JSON válido")
    api_key = body.get('api_key')
    usuario = body.get('usuario')
    clave = body.get('clave')
    ip_usuario = body.get('ip')
    if not all([api_key, usuario, clave, ip_usuario]):
        return _json_response(400, False, "Faltan parámetros (api_key, usuario, clave, ip)")
    # -- 1. Validar API Key del módulo ----------------------------------------
    modulo = API_KEYS_VALIDAS.get(api_key)
    if not modulo:
        return _json_response(401, False, "API Key del módulo inválida o no autorizada")
    # -- 2. Control de Fuerza Bruta (El Guardia) -------------------------------
    try:
        registro = ControlIP.objects.get(ip_usuario=ip_usuario)
        ahora = datetime.now()
        # Aseguramos que ultimo_intento sea naive para la comparación
        ultimo = registro.ultimo_intento
        if hasattr(ultimo, 'tzinfo') and ultimo.tzinfo is not None:
            import pytz
            ultimo = ultimo.astimezone(pytz.utc).replace(tzinfo=None)
        tiempo_pasado = ahora - ultimo
        if registro.intentos_fallidos >= 5:
            if tiempo_pasado < timedelta(minutes=3):
                segundos_restantes = int((timedelta(minutes=3) - tiempo_pasado).total_seconds())
                minutos = segundos_restantes // 60
                segundos = segundos_restantes % 60
                return _json_response(
                    429, False,
                    f"Demasiados intentos. Intente en {minutos}m {segundos}s."
                )
            else:
                # Cumplió el castigo: reiniciar contador
                registro.intentos_fallidos = 0
                registro.save()
    except ControlIP.DoesNotExist:
        pass  # Primera vez que se ve esta IP, todo bien
    # -- 3. Validar credenciales del usuario ----------------------------------
    try:
        user = Usuario.objects.get(user_name=usuario)
    except Usuario.DoesNotExist:
        # No revelamos si el usuario existe o no (seguridad)
        _registrar_intento_fallido(ip_usuario)
        return _json_response(401, False, "Usuario o contraseña incorrectos")
    if not user.estado:
        return _json_response(401, False, "Usuario inactivo")
    es_valido = user.check_password(clave)
    if not es_valido:
        _registrar_intento_fallido(ip_usuario)
        # Pista de auditoría: login fallido
        Auditoria.objects.create(
            username=user,
            accion="LOGIN FALLIDO",
            descripcion=f"Intento fallido desde {ip_usuario}",
            estado_auditoria=False,
            modulo=modulo,
        )
        return _json_response(401, False, "Usuario o contraseña incorrectos")
    # -- 4. Login exitoso: limpiar historial de errores -----------------------
    ControlIP.objects.filter(ip_usuario=ip_usuario).delete()
    # -- 5. Pista de auditoría: login exitoso ---------------------------------
    Auditoria.objects.create(
        username=user,
        accion="LOGIN EXITOSO",
        descripcion=f"Autenticación exitosa desde {ip_usuario}",
        estado_auditoria=True,
        modulo=modulo,
    )
    # -- 6. Generar JWT maestro -----------------------------------------------
    caducidad = datetime.utcnow() + timedelta(hours=8)
    payload = {
        "user_id": user.id,
        "user_name": user.user_name,
        "email": user.email,
        "modulo_origen": modulo,
        "exp": caducidad,
    }
    token = pyjwt.encode(payload, JWT_SECRET, algorithm="HS256")
    return _json_response(200, True, "Autenticación exitosa", token=token)
# -----------------------------------------------------------------------
# Helpers privados
# -----------------------------------------------------------------------
def _registrar_intento_fallido(ip_usuario: str):
    """Incrementa el contador de intentos fallidos para una IP."""
    obj, created = ControlIP.objects.get_or_create(
        ip_usuario=ip_usuario,
        defaults={'intentos_fallidos': 0}
    )
    obj.intentos_fallidos += 1
    obj.ultimo_intento = datetime.now()
    obj.save()
def _json_response(status_code: int, success: bool, message: str, token=None) -> JsonResponse:
    """Helper para devolver respuestas JSON uniformes."""
    body = {'success': success, 'message': message}
    if token:
        body['token'] = token
    return JsonResponse(body, status=status_code)
def roles_view(request):
    roles = Rol.objects.all()
    funciones = Funcion.objects.filter(estado_funcion=True)
    modulos = Modulo.objects.filter(estado_modulo=True).prefetch_related('funciones')
    if request.method == "POST":
        nombre = request.POST.get('nombre_rol', '').strip()
        funciones_seleccionadas = request.POST.getlist('funciones')
        ctx = {'roles': roles, 'funciones': funciones, 'modulos': modulos}
        if not nombre:
            ctx['error'] = 'El nombre del rol es obligatorio.'
            return render(request, 'roles.html', ctx)
        if not funciones_seleccionadas:
            ctx['error'] = 'Debes seleccionar al menos una función para el rol.'
            return render(request, 'roles.html', ctx)
        if Rol.objects.filter(nombre_rol__iexact=nombre).exists():
            ctx['error'] = 'El rol ya existe.'
            return render(request, 'roles.html', ctx)
        rol = Rol.objects.create(nombre_rol=nombre, estado_rol=True)
        for f in funciones_seleccionadas:
            if not FuncionRol.objects.filter(rol=rol, funcion_id=f).exists():
                FuncionRol.objects.create(rol=rol, funcion_id=f)
        return redirect('roles')
    return render(request, 'roles.html', {
        'roles': roles,
        'funciones': funciones,
        'modulos': modulos
    })
@login_required(login_url='login')
def editar_rol(request, id):
    rol = get_object_or_404(Rol, id_rol=id)
    funciones = Funcion.objects.filter(estado_funcion=True)
    modulos = Modulo.objects.filter(estado_modulo=True).prefetch_related('funciones')
    if request.method == "POST":
        nombre = request.POST.get('nombre_rol', '').strip()
        funciones_seleccionadas = request.POST.getlist('funciones')
        ctx = {'rol': rol, 'funciones': funciones, 'modulos': modulos}
        if not nombre:
            ctx['error'] = 'El nombre del rol es obligatorio.'
            return render(request, 'editar_rol.html', ctx)
        if not funciones_seleccionadas:
            ctx['error'] = 'Debes seleccionar al menos una función para el rol.'
            return render(request, 'editar_rol.html', ctx)
        if Rol.objects.filter(nombre_rol__iexact=nombre).exclude(id_rol=id).exists():
            ctx['error'] = 'Ya existe un rol con ese nombre.'
            return render(request, 'editar_rol.html', ctx)
        rol.nombre_rol = nombre
        rol.estado_rol = request.POST.get('estado_rol') in ['on', 'True', 'true', '1']
        rol.save()
        rol.funciones.clear()
        for f in funciones_seleccionadas:
            FuncionRol.objects.create(rol=rol, funcion_id=f)
        return redirect('roles')
    return render(request, 'editar_rol.html', {
        'rol': rol,
        'funciones': funciones,
        'modulos': modulos,
    })
#Usuarios
@login_required(login_url='login')
def usuarios_view(request):
    permiso = Funcion.objects.filter(
        nombre_funcion="SEG_GESTION_USUARIOS",
        roles__usuarios=request.user,
        estado_funcion=True
    ).exists()
    if not permiso:
        return redirect(
            'dashboard_user'
        )
    usuarios = Usuario.objects.all()
    return render(
        request,
        'usuarios.html',
        {
        'usuarios':usuarios
        }
    )
def validar_cedula_ecuatoriana(cedula):
    if len(cedula) != 10 or not cedula.isdigit():
        return False
    provincia = int(cedula[0:2])
    if (provincia < 1 or provincia > 24) and provincia != 30:
        return False
    tercer_digito = int(cedula[2])
    if tercer_digito >= 6:
        return False
    suma = 0
    for i in range(9):
        digito = int(cedula[i])
        if i % 2 == 0:
            digito *= 2
            if digito > 9:
                digito -= 9
        suma += digito
    decena_superior = ((suma // 10) + 1) * 10 if suma % 10 != 0 else suma
    digito_verificador = decena_superior - suma
    if digito_verificador == 10:
        digito_verificador = 0
    return digito_verificador == int(cedula[9])
@login_required(login_url='login')
def crear_usuario(request):
    if not request.user.is_superuser:
        return redirect('usuarios')
    error = None
    if request.method == 'POST':
        user_name = request.POST.get('user_name', '').strip()
        email = request.POST.get('email', '').strip()
        cedula = request.POST.get('cedula', '').strip()
        password = request.POST.get('password')
        estado = request.POST.get('estado') == 'on'
        # Validar Usuario (solo letras)
        if not user_name.isalpha():
            error = "El nombre de usuario solo debe contener letras (sin espacios ni caracteres especiales)."
        # Validar Correo
        elif not (email.endswith('@gmail.com') or email.endswith('@utn.edu.ec')):
            error = "El correo debe ser @gmail.com o @utn.edu.ec."
        # Validar Cédula Ecuatoriana
        elif not validar_cedula_ecuatoriana(cedula):
            error = "La cédula ingresada no es válida en Ecuador."
        # Validar Duplicados
        elif Usuario.objects.filter(user_name=user_name).exists():
            error = f"El usuario '{user_name}' ya existe."
        elif Usuario.objects.filter(email=email).exists():
            error = f"El correo '{email}' ya está registrado."
        elif Usuario.objects.filter(cedula=cedula).exists():
            error = f"La cédula '{cedula}' ya está registrada."
        if not error:
            nuevo_usuario = Usuario.objects.create_user(
                user_name=user_name,
                email=email,
                cedula=cedula,
                password=password
            )
            nuevo_usuario.estado = estado
            nuevo_usuario.save()
            roles_ids = request.POST.getlist('roles')
            if roles_ids:
                nuevo_usuario.roles.set(roles_ids)
            return redirect('usuarios')
        roles_list = Rol.objects.filter(estado_rol=True)
        return render(request, 'crear_usuario.html', {
            'roles_list': roles_list,
            'error': error,
            'old_data': {
                'user_name': user_name,
                'email': email,
                'cedula': cedula,
                'roles': request.POST.getlist('roles')
            }
        })
    roles_list = Rol.objects.filter(estado_rol=True)
    return render(request, 'crear_usuario.html', {
        'roles_list': roles_list,
        'error': error
    })
@login_required(login_url='login')
def editar_usuario(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    error = None
    if request.method == "POST":
        user_name = request.POST.get('user_name', '').strip()
        email = request.POST.get('email', '').strip()
        cedula = request.POST.get('cedula', '').strip()
        password = request.POST.get('password', '').strip()
        estado = request.POST.get('estado') == 'on'
        # Validar Usuario 
        if not user_name.isalpha():
            error = "El nombre de usuario solo debe contener letras (sin espacios ni caracteres especiales)."
        # Validar Correo
        elif not (email.endswith('@gmail.com') or email.endswith('@utn.edu.ec')):
            error = "El correo debe ser @gmail.com o @utn.edu.ec."
        # Validar Cédula Ecuatoriana
        elif not validar_cedula_ecuatoriana(cedula):
            error = "La cédula ingresada no es válida en Ecuador."
        # Validar Duplicados excluyendo el usuario actual
        elif Usuario.objects.filter(user_name=user_name).exclude(id=id).exists():
            error = f"El usuario '{user_name}' ya existe."
        elif Usuario.objects.filter(email=email).exclude(id=id).exists():
            error = f"El correo '{email}' ya está registrado."
        if not error:
            usuario.user_name = user_name
            usuario.email = email
            usuario.cedula = cedula
            usuario.estado = estado
            if password:
                usuario.set_password(password)
            usuario.save()
            roles_ids = request.POST.getlist('roles')
            usuario.roles.clear()
            for rol_id in roles_ids:
                rol = Rol.objects.get(id_rol=rol_id)
                usuario.roles.add(rol)
            return redirect('usuarios')
    roles = Rol.objects.filter(estado_rol=True)
    return render(request, 'editar_usuario.html', {
        'usuario': usuario,
        'roles': roles,
        'error': error
    })
#Funciones
@login_required(login_url='login')
def funciones_view(request):
    permiso = Funcion.objects.filter(
        nombre_funcion="SEG_GESTION_FUNCIONES",
        roles__usuarios=request.user,
        estado_funcion=True
    ).exists()
    if not permiso:
        return redirect('dashboard_user')
    if request.method == "POST":
        nombre = request.POST.get(
            'nombre_funcion'
        )
        estado = True if request.POST.get(
            'estado'
        ) else False
        # VALIDACIÓN DE FUNCIÓN DUPLICADA
        if Funcion.objects.filter(
            nombre_funcion__iexact=nombre
        ).exists():
            funciones = Funcion.objects.all()
            return render(
                request,
                'funciones.html',
                {
                    'funciones': funciones,
                    'error': 'La función ya existe'
                }
            )
        Funcion.objects.create(
            nombre_funcion=nombre,
            estado_funcion=estado
        )
        return redirect('funciones')
    funciones = Funcion.objects.all()
    roles = Rol.objects.all()
    from modules.models import Modulo
    modulos = Modulo.objects.all()
    return render(
        request,
        'funciones.html',
        {
            'funciones': funciones,
            'roles': roles,
            'modulos': modulos
        }
    )
@login_required(login_url='login')
def editar_funcion(request,id):
    funcion = get_object_or_404(
        Funcion,
        id_funcion=id
    )
    if request.method == "POST":
        funcion.nombre_funcion = request.POST.get(
            'nombre_funcion'
        )
        funcion.estado_funcion = True if request.POST.get(
            'estado'
        ) else False
        funcion.save()
        return redirect(
            'funciones'
        )
    return render(
        request,
        'editar_funcion.html',
        {
        'funcion':funcion
        }
    )
#Login
def login_view(request):
    if request.user.is_authenticated:
        is_admin = request.user.roles.filter(nombre_rol__icontains='admin').exists() or request.user.roles.filter(nombre_rol__icontains='seguridad').exists() or request.user.is_superuser
        return redirect('dashboard_admin' if is_admin else 'dashboard_user')
    error = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            user = Usuario.objects.get(user_name=username)
            if not user.estado:
                error = "El usuario está inactivo (borrado lógico)."
            else:
                # Obtener IP del cliente
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                ip_usuario = x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR')
                # Primero intentamos validación local de contraseña
                if user.check_password(password):
                    is_valid = True
                    ms_result = "Autenticación Local"
                else:
                    is_valid, ms_result = validate_microsoft_token(username, password)
                if is_valid:
                    # Asignar backend explícitamente para iniciar sesión en la sesión de Django
                    user.backend = 'django.contrib.auth.backends.ModelBackend'
                    auth_login(request, user)
                    # Registrar auditoría de éxito
                    Auditoria.objects.create(
                        username=user,
                        accion="LOGIN EXITOSO (WEB)",
                        descripcion="Acceso web correcto desde portal unificado",
                        observacion=f"IP: {ip_usuario}",
                        estado_auditoria=True
                    )
                    # Generar JWT propio para consumo
                    token = generate_jwt(user)
                    # Determinar si el usuario posee roles administrativos
                    is_admin = user.roles.filter(nombre_rol__icontains='admin').exists() or user.roles.filter(nombre_rol__icontains='seguridad').exists() or user.is_superuser
                    response = redirect('dashboard_admin' if is_admin else 'dashboard_user')
                    # Guardar token JWT en cookies para poder leerlo desde el cliente/API
                    response.set_cookie('jwt_token', token, max_age=7200, httponly=False)
                    return response
                else:
                    # Registrar auditoría de fallo
                    Auditoria.objects.create(
                        username=user,
                        accion="LOGIN FALLIDO (WEB)",
                        descripcion=f"Intento web fallido: {ms_result}",
                        observacion=f"IP: {ip_usuario}",
                        estado_auditoria=False
                    )
                    error = f"Contraseña incorrecta o error de autenticación: {ms_result}"
        except Usuario.DoesNotExist:
            error = "El usuario no está registrado en el Módulo de Seguridad."
    return render(request, 'login.html', {'error': error})
#Modulos
from modules.models import Modulo
@login_required(login_url='login')
def modulos_view(request):
    if request.method == "POST":
        action = request.POST.get('action')
        if action == 'create_module':
            nombre = request.POST.get('nombre_modulo')
            descripcion = request.POST.get('descripcion_modulo')
            # estado por defecto True en BD
            try:
                nuevo_mod = Modulo(nombre_modulo=nombre, descripcion_modulo=descripcion)
                nuevo_mod.full_clean()
                nuevo_mod.save()
            except Exception as e:
                request.session['modulo_error'] = "Error al crear el módulo. Verifica que el nombre no esté duplicado."
        elif action == 'toggle_module_status':
            mod_id = request.POST.get('module_id')
            try:
                mod = Modulo.objects.get(id_modulo=mod_id)
                mod.estado_modulo = not mod.estado_modulo
                mod.save()
            except Modulo.DoesNotExist:
                pass
        return redirect('modulos')
    modulos = Modulo.objects.all().order_by('-id_modulo')
    error = request.session.pop('modulo_error', None)
    return render(
        request,
        'modulos.html',
        {
            'modulos': modulos,
            'error': error
        }
    )
def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard_user')
    error = None
    success = None
    if request.method == 'POST':
        import re
        from django.db import IntegrityError
        username = request.POST.get('username')
        email = request.POST.get('email')
        cedula = request.POST.get('cedula')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        if password != confirm_password:
            error = "Las contraseñas no coinciden."
        elif not re.match(r'^[a-zA-Z0-9._%+-]+@(gmail\.com|utn\.edu\.ec)$', email):
            error = "Solo se permiten correos @gmail.com o @utn.edu.ec."
        elif not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*[@$!%*?&\.\-\_])[A-Za-z\d@$!%*?&\.\-\_]{8,}$', password):
            error = "La contraseña debe tener 8+ caracteres, mayúscula, minúscula y un carácter especial (@$!%*?&.-_)."
        elif Usuario.objects.filter(user_name=username).exists():
            error = "El nombre de usuario ya está registrado."
        elif Usuario.objects.filter(email=email).exists():
            error = "El correo electrónico ya está registrado."
        else:
            try:
                user = Usuario.objects.create_user(
                    user_name=username,
                    email=email,
                    cedula=cedula,
                    password=password
                )
                from users.models import EmailVerificationToken
                from django.core.mail import send_mail
                from django.conf import settings
                token = EmailVerificationToken.objects.create(usuario=user)
                verify_url = request.build_absolute_uri(reverse('verify_email', args=[str(token.token)]))
                try:
                    send_mail(
                        subject='Bienvenido - Verifica tu correo',
                        message=f'Hola {user.user_name},\n\nGracias por registrarte en Seguridad Centralizada. Por favor verifica tu correo electrónico haciendo clic en el siguiente enlace:\n{verify_url}',
                        from_email=settings.EMAIL_HOST_USER,
                        recipient_list=[email],
                        fail_silently=False,
                    )
                except Exception:
                    pass # Evitamos romper el flujo si el correo falla
                success = "Usuario registrado con éxito. Revisa tu correo electrónico para verificar tu cuenta e iniciar sesión."
                # Al tener éxito, renderizamos login.html sin el flag show_register,
                # para que se posicione directamente en la pestaña de Iniciar Sesión.
                return render(request, 'login.html', {
                    'register_success': success
                })
            except IntegrityError as e:
                error_msg = str(e).lower()
                if 'cedula_valida' in error_msg:
                    error = "La cédula ingresada no es válida en Ecuador."
                elif 'cedula' in error_msg:
                    error = "Esta cédula ya se encuentra registrada."
                else:
                    error = "Error de base de datos: Verifica tus datos."
            except Exception as e:
                error = f"Error general al registrar: {str(e)}"
    return render(request, 'login.html', {
        'register_error': error, 
        'show_register': True
    })
def logout_view(request):
    auth_logout(request)
    response = redirect('login')
    response.delete_cookie('jwt_token')
    return response
def custom_csrf_failure(request, reason=""):
    return render(request, 'login.html', {
        'register_error': 'La solicitud ha expirado o es inválida por seguridad (CSRF). Por favor, intenta de nuevo.',
        'show_forgot': True
    })
def forgot_password_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = Usuario.objects.get(email=email)
            PasswordResetToken.objects.filter(usuario=user, usado=False).update(usado=True)
            token = PasswordResetToken.objects.create(usuario=user, codigo=generar_codigo_6_digitos())
            try:
                send_mail(
                    subject='Código de Recuperación - Seguridad Centralizada',
                    message=f'Hola {user.user_name},\n\nTu código de recuperación de 6 dígitos es:\n{token.codigo}\n\nIngresa este código en la aplicación para restablecer tu contraseña. Si no solicitaste esto, ignora este correo.',
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[email],
                    fail_silently=False,
                )
                success = "Te hemos enviado un código de 6 dígitos a tu correo."
                return render(request, 'verify_code.html', {'email': email, 'success': success})
            except Exception as e:
                success = f"Simulación local (envío de email fallido). Tu código es: {token.codigo}"
                return render(request, 'verify_code.html', {'email': email, 'success': success})
        except Usuario.DoesNotExist:
            error = "No existe ningún usuario con este correo institucional."
            return render(request, 'login.html', {'register_error': error, 'show_forgot': True})
    return redirect('login')
def verify_reset_code_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        codigo = request.POST.get('codigo')
        if codigo:
            codigo = codigo.upper().strip()
        try:
            user = Usuario.objects.get(email=email)
            reset_token = PasswordResetToken.objects.filter(usuario=user, codigo=codigo, usado=False).first()
            if reset_token:
                request.session['reset_authorized_email'] = email
                request.session['reset_token_id'] = reset_token.id
                return redirect('reset_password')
            else:
                return render(request, 'verify_code.html', {'email': email, 'error': 'El código es inválido o ya expiró.'})
        except Usuario.DoesNotExist:
            return render(request, 'login.html', {'register_error': 'Usuario no válido.', 'show_forgot': True})
    return redirect('login')
def verify_email_view(request, token):
    from users.models import EmailVerificationToken
    try:
        verification_token = EmailVerificationToken.objects.get(token=token, usado=False)
        user = verification_token.usuario
        user.correo_verificado = True
        user.save()
        verification_token.usado = True
        verification_token.save()
        return render(request, 'login.html', {
            'register_success': '¡Correo verificado exitosamente! Ya puedes iniciar sesión.'
        })
    except EmailVerificationToken.DoesNotExist:
        return render(request, 'login.html', {
            'error': 'El enlace de verificación es inválido o ya fue utilizado.'
        })
def reset_password_view(request):
    email = request.session.get('reset_authorized_email')
    token_id = request.session.get('reset_token_id')
    if not email or not token_id:
        return redirect('login')
    try:
        reset_token = PasswordResetToken.objects.get(id=token_id, usado=False)
    except PasswordResetToken.DoesNotExist:
        return redirect('login')
    if request.method == 'POST':
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        if password != confirm_password:
            return render(request, 'reset_password.html', {'error': 'Las contraseñas no coinciden.'})
        if len(password) < 8:
            return render(request, 'reset_password.html', {'error': 'La contraseña debe tener al menos 8 caracteres.'})
        user = reset_token.usuario
        user.set_password(password)
        user.save()
        reset_token.usado = True
        reset_token.save()
        if 'reset_authorized_email' in request.session:
            del request.session['reset_authorized_email']
        if 'reset_token_id' in request.session:
            del request.session['reset_token_id']
        return render(request, 'login.html', {'register_success': 'Contraseña actualizada con éxito. Ahora puedes iniciar sesión.'})
    return render(request, 'reset_password.html')
@login_required(login_url='login')
def dashboard_user_view(request):
    token = request.COOKIES.get('jwt_token')
    return render(request, 'dashboard_user.html', {
        'user': request.user,
        'token': token
    })
@login_required(login_url='login')
def dashboard_admin_view(request):
    # Validar si el usuario tiene rol administrativo
    is_admin = (
        request.user.roles.filter(nombre_rol__icontains='admin').exists()
        or request.user.roles.filter(nombre_rol__icontains='seguridad').exists()
        or request.user.is_superuser
    )
    if not is_admin:
        return redirect('dashboard_user')
    token = request.COOKIES.get('jwt_token')
    # ==========================
    # MANEJO DE ACCIONES POST
    # ==========================
    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = request.POST.get('user_id')
        if action == 'toggle_status' and user_id:
            try:
                target_user = Usuario.objects.get(id=user_id)
                if target_user.id != request.user.id:
                    target_user.estado = not target_user.estado
                    target_user.save()
            except Usuario.DoesNotExist:
                pass
        elif action == 'edit_roles' and user_id:
            try:
                target_user = Usuario.objects.get(id=user_id)
                roles_ids = request.POST.getlist('roles')
                target_user.roles.set(roles_ids)
                target_user.save()
            except Usuario.DoesNotExist:
                pass
        return redirect('/dashboard/admin/?tab=users')
    # ==========================
    # ESTADÍSTICAS
    # ==========================
    stats = {
        'total_users': Usuario.objects.count(),
        'active_users': Usuario.objects.filter(
            estado=True
        ).count(),
        'total_roles': Rol.objects.count(),
        'total_modules': Modulo.objects.count(),
        'total_logs': Auditoria.objects.count()
    }
    # ==========================
    # AUDITORÍA (LOGS)
    # ==========================
    search_log = request.GET.get('search_log', '')
    date_log = request.GET.get('date_log', '')
    order_log = request.GET.get('order_log', 'desc')
    logs_query = Auditoria.objects.all()
    if search_log:
        logs_query = logs_query.filter(
            Q(username__user_name__icontains=search_log) |
            Q(observacion__icontains=search_log)
        )
    if date_log:
        logs_query = logs_query.filter(fecha_creacion__date=date_log)
    if order_log == 'asc':
        logs_query = logs_query.order_by('fecha_creacion')
    else:
        logs_query = logs_query.order_by('-fecha_creacion')
    logs = logs_query[:1000]  # Limitar a 1000 para no saturar, la UI paginará esto
    # ==========================
    # FUNCIONES DEL USUARIO
    # ==========================
    funciones = Funcion.objects.filter(
        roles__usuarios=request.user,
        estado_funcion=True
    ).distinct()
    # ==========================
    # USUARIOS
    # ==========================
    search_query = request.GET.get('q', '')
    users_list = Usuario.objects.all() \
        .prefetch_related('roles') \
        .order_by('-fecha_creacion')
    if search_query:
        users_list = users_list.filter(
            Q(user_name__icontains=search_query)
            |
            Q(email__icontains=search_query)
            |
            Q(cedula__icontains=search_query)
        )
    paginator = Paginator(
        users_list,
        10
    )
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(
        page_number
    )
    # ==========================
    # ROLES
    # ==========================
    roles_list = Rol.objects.all()
    # ==========================
    # TAB ACTIVA
    # ==========================
    active_tab = request.GET.get(
        'tab',
        'dashboard'
    )
    return render(
        request,
        'dashboard_admin.html',
        {
            'user': request.user,
            'token': token,
            'stats': stats,
            'logs': logs,
            'funciones': funciones,
            'page_obj': page_obj,
            'roles_list': roles_list,
            'search_query': search_query,
            'active_tab': active_tab
        }
    )
@login_required(login_url='login')
def gestion_view(request):
    permiso = request.user.is_superuser or Funcion.objects.filter(
        nombre_funcion__in=["SEG_GESTION_USUARIOS", "SEG_GESTION_ROLES", "SEG_GESTION_MODULOS"],
        roles__usuarios=request.user,
        estado_funcion=True
    ).exists()
    if not permiso:
        return redirect('dashboard_user')
    usuarios = Usuario.objects.all()
    roles = Rol.objects.all()
    modulos = Modulo.objects.all()
    funciones = Funcion.objects.all()
    return render(request, 'gestion.html', {
        'usuarios': usuarios,
        'roles': roles,
        'modulos': modulos,
        'funciones': funciones
    })
@csrf_exempt
@require_POST
def api_forgot_password(request):
    try:
        body = json.loads(request.body)
        email = body.get('email')
        if not email:
            return _json_response(400, False, 'Falta email')
        user = Usuario.objects.get(email=email)
        PasswordResetToken.objects.filter(usuario=user, usado=False).update(usado=True)
        token = PasswordResetToken.objects.create(usuario=user, codigo=generar_codigo_6_digitos())
        try:
            send_mail(
                subject='Código de Recuperación - API',
                message=f'Hola {user.user_name},\n\nTu código de recuperación es:\n{token.codigo}',
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[email],
                fail_silently=False,
            )
            return _json_response(200, True, 'Código enviado al correo')
        except Exception as e:
            return _json_response(200, True, f'Simulación local. Tu código es: {token.codigo}')
    except Usuario.DoesNotExist:
        return _json_response(404, False, 'Usuario no encontrado')
    except Exception as e:
        return _json_response(500, False, str(e))
@csrf_exempt
@require_POST
def api_verify_code(request):
    try:
        body = json.loads(request.body)
        email = body.get('email')
        codigo = body.get('codigo')
        if not email or not codigo:
            return _json_response(400, False, 'Faltan parámetros')
        codigo = codigo.upper().strip()
        user = Usuario.objects.get(email=email)
        token = PasswordResetToken.objects.filter(usuario=user, codigo=codigo, usado=False).first()
        if token:
            return _json_response(200, True, 'Código válido')
        else:
            return _json_response(400, False, 'Código inválido o expirado')
    except Usuario.DoesNotExist:
        return _json_response(404, False, 'Usuario no encontrado')
    except Exception as e:
        return _json_response(500, False, str(e))
@csrf_exempt
@require_POST
def api_reset_password(request):
    try:
        body = json.loads(request.body)
        email = body.get('email')
        codigo = body.get('codigo')
        new_password = body.get('new_password')
        if not email or not codigo or not new_password:
            return _json_response(400, False, 'Faltan parámetros')
        codigo = codigo.upper().strip()
        user = Usuario.objects.get(email=email)
        token = PasswordResetToken.objects.filter(usuario=user, codigo=codigo, usado=False).first()
        if not token:
            return _json_response(400, False, 'Código inválido o expirado')
        if len(new_password) < 8:
            return _json_response(400, False, 'La contraseña debe tener al menos 8 caracteres')
        user.set_password(new_password)
        user.save()
        token.usado = True
        token.save()
        return _json_response(200, True, 'Contraseña actualizada con éxito')
    except Usuario.DoesNotExist:
        return _json_response(404, False, 'Usuario no encontrado')
    except Exception as e:
        return _json_response(500, False, str(e))
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
@login_required(login_url='login')
def exportar_reporte_excel(request, tipo):
    if not (request.user.roles.filter(nombre_rol__icontains='admin').exists() or request.user.roles.filter(nombre_rol__icontains='seguridad').exists() or request.user.is_superuser):
        return redirect('dashboard_user')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Reporte_{tipo.capitalize()}'
    # Estilos cabecera
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    if tipo == 'usuarios':
        ws.append(['ID', 'Username', 'Email', 'Cdula', 'Estado', 'Roles'])
        for usuario in Usuario.objects.all().prefetch_related('roles'):
            roles = ', '.join([r.nombre_rol for r in usuario.roles.all()])
            estado = 'Activo' if usuario.estado else 'Inactivo'
            ws.append([usuario.id, usuario.user_name, usuario.email, usuario.cedula, estado, roles])
    elif tipo == 'roles':
        ws.append(['ID', 'Nombre Rol', 'Estado', 'Funciones Asignadas'])
        for rol in Rol.objects.all().prefetch_related('funciones'):
            funciones = ', '.join([f.nombre_funcion for f in rol.funciones.all()])
            estado = 'Activo' if rol.estado_rol else 'Inactivo'
            ws.append([rol.id_rol, rol.nombre_rol, estado, funciones])
    elif tipo == 'modulos':
        ws.append(['ID', 'Nombre Módulo', 'Descripción', 'Estado'])
        for modulo in Modulo.objects.all():
            estado = 'Activo' if modulo.estado_modulo else 'Inactivo'
            ws.append([modulo.id_modulo, modulo.nombre_modulo, modulo.descripcion_modulo, estado])
    elif tipo == 'funciones':
        ws.append(['ID', 'Nombre Funcin', 'Estado'])
        for funcion in Funcion.objects.all():
            estado = 'Activo' if funcion.estado_funcion else 'Inactivo'
            ws.append([funcion.id_funcion, funcion.nombre_funcion, estado])
    elif tipo == 'auditoria':
        ws.append(['ID', 'Usuario', 'Accin', 'Descripcin', 'Observacin', 'Estado', 'Fecha'])
        for log in Auditoria.objects.all().order_by('-fecha_creacion'):
            estado = 'xito' if log.estado_auditoria else 'Error'
            fecha = log.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if log.fecha_creacion else ''
            ws.append([log.id_auditoria, str(log.username), log.accion, log.descripcion, log.observacion, estado, fecha])
    else:
        return HttpResponse('Tipo de reporte no vlido')
    # Dar formato a las cabeceras
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    # Ajustar ancho de columnas automticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_{tipo.capitalize()}.xlsx"'
    return response
